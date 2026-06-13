"""
═══════════════════════════════════════════════════════════════
  CA FirmHub — SaaS / Cloud Version
  Multi-tenant Flask + PostgreSQL (Neon.tech)
  Deploy on Render.com (free tier)
═══════════════════════════════════════════════════════════════
"""
import os, uuid, json, re
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, g, send_file, send_from_directory
import psycopg2
import psycopg2.extras
import jwt
from werkzeug.security import generate_password_hash, check_password_hash

# ── Config from environment variables ──────────────────────────────────
DATABASE_URL   = os.environ.get("DATABASE_URL")          # Neon PostgreSQL URL
SECRET_KEY     = os.environ.get("SECRET_KEY", "change-me-in-production")
SUPERADMIN_KEY = os.environ.get("SUPERADMIN_KEY", "superadmin-secret-key")  # for /superadmin
TOKEN_HOURS    = int(os.environ.get("TOKEN_HOURS", 12))
PORT           = int(os.environ.get("PORT", 8000))

# ── Cloudflare R2 (optional file storage) ──────────────────────────────
R2_ACCOUNT_ID  = os.environ.get("R2_ACCOUNT_ID", "")
R2_ACCESS_KEY  = os.environ.get("R2_ACCESS_KEY", "")
R2_SECRET_KEY  = os.environ.get("R2_SECRET_KEY", "")
R2_BUCKET      = os.environ.get("R2_BUCKET", "cafirmhub-files")

app = Flask(__name__, static_folder="static")
app.config["SECRET_KEY"] = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# ═══════════════════════════════════════════════════════════════
#  DATABASE — PostgreSQL via Neon.tech
# ═══════════════════════════════════════════════════════════════
def get_db():
    if "db" not in g:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = False
        g.db = conn
    return g.db

@app.teardown_appcontext
def close_db(exc=None):
    db = g.pop("db", None)
    if db:
        if exc:
            db.rollback()
        db.close()

def qry(sql, params=(), one=False, commit=False):
    """Execute SQL, return dict rows."""
    db = get_db()
    cur = db.cursor()
    # Convert SQLite ? placeholders → PostgreSQL %s
    pg_sql = sql.replace("?", "%s")
    cur.execute(pg_sql, params)
    if commit:
        db.commit()
    if one:
        row = cur.fetchone()
        return dict(row) if row else None
    rows = cur.fetchall()
    return [dict(r) for r in rows]

def qry_id(sql, params=()):
    """Execute INSERT and return the new row id (PostgreSQL RETURNING)."""
    db = get_db()
    cur = db.cursor()
    pg_sql = sql.replace("?", "%s")
    # Append RETURNING id if not present
    if "RETURNING" not in pg_sql.upper():
        pg_sql += " RETURNING id"
    cur.execute(pg_sql, params)
    db.commit()
    row = cur.fetchone()
    return row["id"] if row else None

def execute(sql, params=(), commit=True):
    """Execute a statement (UPDATE/DELETE/INSERT without return)."""
    db = get_db()
    cur = db.cursor()
    pg_sql = sql.replace("?", "%s")
    cur.execute(pg_sql, params)
    if commit:
        db.commit()

# ── Init all tables ─────────────────────────────────────────────────────
def init_db():
    db = get_db()
    cur = db.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS firms (
        id SERIAL PRIMARY KEY,
        name TEXT NOT NULL,
        reg_no TEXT,
        sub_id TEXT UNIQUE NOT NULL,
        is_active BOOLEAN DEFAULT TRUE,
        expires_at DATE,
        activated BOOLEAN DEFAULT FALSE,
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        username TEXT NOT NULL,
        full_name TEXT NOT NULL,
        email TEXT,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'Member'
            CHECK(role IN ('Admin','Team Leader','Member','Client')),
        is_active BOOLEAN DEFAULT TRUE,
        client_id INTEGER,
        created_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(firm_id, username)
    );

    CREATE TABLE IF NOT EXISTS clients (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        name TEXT NOT NULL, pan TEXT, gstin TEXT, address TEXT,
        contact_person TEXT, contact_phone TEXT, contact_email TEXT,
        is_active BOOLEAN DEFAULT TRUE,
        created_by_id INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS engagements (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        client_id INTEGER NOT NULL REFERENCES clients(id),
        title TEXT NOT NULL, engagement_type TEXT NOT NULL,
        financial_year TEXT NOT NULL, period_from DATE, period_to DATE,
        team_leader_id INTEGER REFERENCES users(id),
        status TEXT DEFAULT 'Active', notes TEXT,
        created_by_id INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS engagement_teams (
        id SERIAL PRIMARY KEY,
        engagement_id INTEGER NOT NULL REFERENCES engagements(id) ON DELETE CASCADE,
        user_id INTEGER NOT NULL REFERENCES users(id),
        role_in_engagement TEXT DEFAULT 'Member'
    );

    CREATE TABLE IF NOT EXISTS audit_programs (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        name TEXT NOT NULL, engagement_type TEXT NOT NULL, description TEXT,
        is_active BOOLEAN DEFAULT TRUE,
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS audit_checklist_items (
        id SERIAL PRIMARY KEY,
        program_id INTEGER NOT NULL REFERENCES audit_programs(id) ON DELETE CASCADE,
        sr_no INTEGER NOT NULL, area TEXT NOT NULL, description TEXT NOT NULL,
        reference TEXT, priority TEXT DEFAULT 'Medium'
    );

    CREATE TABLE IF NOT EXISTS tasks (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        engagement_id INTEGER NOT NULL REFERENCES engagements(id),
        checklist_item_id INTEGER REFERENCES audit_checklist_items(id),
        title TEXT NOT NULL, description TEXT, area TEXT,
        assignee_id INTEGER REFERENCES users(id),
        status TEXT DEFAULT 'Pending'
            CHECK(status IN ('Pending','In Progress','Completed','Under Review','Approved','Rejected')),
        priority TEXT DEFAULT 'Medium', due_date DATE, completed_at TIMESTAMP,
        working_paper_ref TEXT, estimated_hours REAL,
        created_by_id INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS task_assignees (
        task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
        assigned_at TIMESTAMP DEFAULT NOW(),
        PRIMARY KEY (task_id, user_id)
    );

    CREATE TABLE IF NOT EXISTS comments (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        author_id INTEGER NOT NULL REFERENCES users(id),
        content TEXT NOT NULL, is_query BOOLEAN DEFAULT FALSE,
        created_at TIMESTAMP DEFAULT NOW(),
        updated_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS queries (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        engagement_id INTEGER NOT NULL REFERENCES engagements(id),
        comment_id INTEGER REFERENCES comments(id),
        sr_no INTEGER NOT NULL, query_text TEXT NOT NULL, response TEXT,
        status TEXT DEFAULT 'Open' CHECK(status IN ('Open','Responded','Closed')),
        raised_by_id INTEGER NOT NULL REFERENCES users(id),
        raised_date TIMESTAMP DEFAULT NOW(),
        responded_by_id INTEGER REFERENCES users(id),
        responded_date TIMESTAMP, task_reference TEXT
    );

    CREATE TABLE IF NOT EXISTS file_uploads (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        filename TEXT NOT NULL, original_filename TEXT NOT NULL,
        file_key TEXT NOT NULL,
        file_size INTEGER, mime_type TEXT,
        uploaded_by_id INTEGER NOT NULL REFERENCES users(id),
        uploaded_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS reviews (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        reviewer_id INTEGER NOT NULL REFERENCES users(id),
        action TEXT NOT NULL CHECK(action IN ('Approved','Rejected')),
        remarks TEXT, reviewed_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS time_logs (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
        user_id INTEGER NOT NULL REFERENCES users(id),
        date DATE NOT NULL, hours REAL NOT NULL, note TEXT,
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS invoices (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        engagement_id INTEGER NOT NULL REFERENCES engagements(id),
        client_id INTEGER NOT NULL REFERENCES clients(id),
        invoice_no TEXT NOT NULL,
        invoice_date DATE NOT NULL,
        description TEXT,
        amount REAL NOT NULL,
        gst_percent REAL DEFAULT 18.0,
        gst_amount REAL DEFAULT 0,
        total_amount REAL NOT NULL,
        payment_status TEXT DEFAULT 'Unpaid'
            CHECK(payment_status IN ('Unpaid','Partial','Paid')),
        payment_date DATE, payment_note TEXT,
        created_by_id INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS doc_register (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        engagement_id INTEGER REFERENCES engagements(id),
        client_id INTEGER REFERENCES clients(id),
        doc_type TEXT NOT NULL CHECK(doc_type IN ('Inward','Outward')),
        doc_name TEXT NOT NULL, doc_category TEXT, doc_date DATE NOT NULL,
        received_from TEXT, sent_to TEXT, reference_no TEXT,
        status TEXT DEFAULT 'Received'
            CHECK(status IN ('Received','Acknowledged','Dispatched','Returned','Pending')),
        remarks TEXT,
        created_by_id INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS compliance_calendar (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER NOT NULL REFERENCES firms(id) ON DELETE CASCADE,
        title TEXT NOT NULL, category TEXT NOT NULL, due_date DATE NOT NULL,
        description TEXT, financial_year TEXT,
        engagement_id INTEGER REFERENCES engagements(id),
        is_recurring BOOLEAN DEFAULT FALSE,
        status TEXT DEFAULT 'Upcoming'
            CHECK(status IN ('Upcoming','Completed','Missed')),
        created_by_id INTEGER REFERENCES users(id),
        created_at TIMESTAMP DEFAULT NOW()
    );

    CREATE TABLE IF NOT EXISTS audit_logs (
        id SERIAL PRIMARY KEY,
        firm_id INTEGER REFERENCES firms(id) ON DELETE SET NULL,
        user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
        action TEXT NOT NULL, entity_type TEXT, entity_id INTEGER,
        details TEXT, ip_address TEXT,
        timestamp TIMESTAMP DEFAULT NOW()
    );

    -- Indexes
    CREATE INDEX IF NOT EXISTS idx_tasks_firm        ON tasks(firm_id);
    CREATE INDEX IF NOT EXISTS idx_tasks_engagement  ON tasks(engagement_id);
    CREATE INDEX IF NOT EXISTS idx_tasks_status      ON tasks(status);
    CREATE INDEX IF NOT EXISTS idx_comments_task     ON comments(task_id);
    CREATE INDEX IF NOT EXISTS idx_queries_eng       ON queries(engagement_id);
    CREATE INDEX IF NOT EXISTS idx_invoices_firm     ON invoices(firm_id);
    CREATE INDEX IF NOT EXISTS idx_compliance_due    ON compliance_calendar(due_date);
    """)
    db.commit()
    print("[DB] Tables initialised.")

# ═══════════════════════════════════════════════════════════════
#  AUTH HELPERS
# ═══════════════════════════════════════════════════════════════
def hash_pw(pw):   return generate_password_hash(pw)
def check_pw(pw, h): return check_password_hash(h, pw)

def make_token(user_id, firm_id, role):
    payload = {
        "sub": str(user_id),
        "firm": firm_id,
        "role": role,
        "exp": datetime.utcnow() + timedelta(hours=TOKEN_HOURS)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"detail": "Not authenticated"}), 401
        try:
            payload = jwt.decode(auth.split(" ", 1)[1], SECRET_KEY, algorithms=["HS256"])
        except jwt.ExpiredSignatureError:
            return jsonify({"detail": "Token expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"detail": "Invalid token"}), 401
        user = qry("SELECT * FROM users WHERE id=%s AND is_active=TRUE",
                   (int(payload["sub"]),), one=True)
        if not user:
            return jsonify({"detail": "User not found"}), 401
        g.user = user
        g.firm_id = user["firm_id"]
        return f(*args, **kwargs)
    return wrapper

def require_role(*roles):
    def decorator(f):
        @wraps(f)
        @login_required
        def wrapper(*args, **kwargs):
            if g.user["role"] not in roles:
                return jsonify({"detail": "Insufficient permissions"}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator

def superadmin_required(f):
    """Require X-SuperAdmin-Key header matching SUPERADMIN_KEY."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        key = request.headers.get("X-SuperAdmin-Key", "")
        if key != SUPERADMIN_KEY:
            return jsonify({"detail": "Super-admin access required"}), 403
        return f(*args, **kwargs)
    return wrapper

def log_action(firm_id, user_id, action, entity_type=None, entity_id=None, details=None, ip=None):
    execute(
        "INSERT INTO audit_logs (firm_id,user_id,action,entity_type,entity_id,details,ip_address) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (firm_id, user_id, action, entity_type, entity_id, details, ip),
        commit=True
    )

def _user_out(u):
    return {k: u[k] for k in
            ["id","username","full_name","email","role","is_active","client_id","created_at","firm_id"]}

# ═══════════════════════════════════════════════════════════════
#  SERVE FRONTEND
# ═══════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/superadmin")
def superadmin_page():
    return send_from_directory("static", "superadmin.html")

@app.route("/tools/workpapers")
def workpapers_page():
    return send_from_directory("static", "audit_working_papers.html")
@app.route("/static/<path:fn>")
def static_files(fn):
    return send_from_directory("static", fn)

# ═══════════════════════════════════════════════════════════════
#  SUPER-ADMIN PANEL API  (protected by SUPERADMIN_KEY header)
# ═══════════════════════════════════════════════════════════════

@app.route("/api/superadmin/firms", methods=["GET"])
@superadmin_required
def sa_list_firms():
    firms = qry("SELECT * FROM firms ORDER BY created_at DESC")
    for f in firms:
        f["user_count"] = qry("SELECT COUNT(*) as c FROM users WHERE firm_id=%s",
                               (f["id"],), one=True)["c"]
    return jsonify(firms)

@app.route("/api/superadmin/firms", methods=["POST"])
@superadmin_required
def sa_create_firm():
    d = request.get_json()
    name    = d.get("name","").strip()
    reg_no  = d.get("reg_no","").strip()
    expires = d.get("expires_at")        # "YYYY-MM-DD" or None
    if not name:
        return jsonify({"detail": "Firm name required"}), 400
    sub_id = "SUB-" + uuid.uuid4().hex[:12].upper()
    fid = qry_id(
        "INSERT INTO firms (name,reg_no,sub_id,is_active,expires_at) VALUES (%s,%s,%s,TRUE,%s)",
        (name, reg_no, sub_id, expires)
    )
    return jsonify({"id": fid, "name": name, "sub_id": sub_id,
                    "reg_no": reg_no, "expires_at": expires}), 201

@app.route("/api/superadmin/firms/<int:fid>", methods=["PUT"])
@superadmin_required
def sa_update_firm(fid):
    d = request.get_json()
    fields, vals = [], []
    for col in ["name","reg_no","expires_at","is_active"]:
        if col in d:
            fields.append(f"{col}=%s")
            vals.append(d[col])
    if not fields:
        return jsonify({"detail":"Nothing to update"}), 400
    vals.append(fid)
    execute(f"UPDATE firms SET {','.join(fields)} WHERE id=%s", vals)
    return jsonify(qry("SELECT * FROM firms WHERE id=%s", (fid,), one=True))

@app.route("/api/superadmin/firms/<int:fid>/users", methods=["GET"])
@superadmin_required
def sa_firm_users(fid):
    return jsonify([_user_out(u) for u in
                    qry("SELECT * FROM users WHERE firm_id=%s ORDER BY full_name", (fid,))])

@app.route("/api/superadmin/firms/<int:fid>/users", methods=["POST"])
@superadmin_required
def sa_create_user(fid):
    d = request.get_json()
    username = d.get("username","").strip()
    if not username:
        return jsonify({"detail": "Username required"}), 400
    exists = qry("SELECT id FROM users WHERE firm_id=%s AND username=%s",
                 (fid, username), one=True)
    if exists:
        return jsonify({"detail": "Username already exists in this firm"}), 400
    uid = qry_id(
        "INSERT INTO users (firm_id,username,full_name,email,password_hash,role) "
        "VALUES (%s,%s,%s,%s,%s,%s)",
        (fid, username, d.get("full_name",""), d.get("email",""),
         hash_pw(d.get("password","audit123")), d.get("role","Admin"))
    )
    return jsonify(_user_out(qry("SELECT * FROM users WHERE id=%s", (uid,), one=True))), 201

@app.route("/api/superadmin/stats", methods=["GET"])
@superadmin_required
def sa_stats():
    return jsonify({
        "total_firms":   qry("SELECT COUNT(*) as c FROM firms",          one=True)["c"],
        "active_firms":  qry("SELECT COUNT(*) as c FROM firms WHERE is_active=TRUE", one=True)["c"],
        "total_users":   qry("SELECT COUNT(*) as c FROM users",          one=True)["c"],
        "total_clients": qry("SELECT COUNT(*) as c FROM clients",        one=True)["c"],
        "total_tasks":   qry("SELECT COUNT(*) as c FROM tasks",          one=True)["c"],
    })

# ── Regenerate Subscription ID ─────────────────────────────────
@app.route("/api/superadmin/firms/<int:fid>/regen-sub", methods=["POST"])
@superadmin_required
def sa_regen_sub(fid):
    new_sub = "SUB-" + uuid.uuid4().hex[:12].upper()
    execute("UPDATE firms SET sub_id=%s, activated=FALSE WHERE id=%s", (new_sub, fid))
    return jsonify({"sub_id": new_sub})

# ═══════════════════════════════════════════════════════════════
#  SUBSCRIPTION / ACTIVATION
# ═══════════════════════════════════════════════════════════════
@app.route("/api/subscription/verify", methods=["POST"])
def subscription_verify():
    d = request.get_json(silent=True) or {}
    sub_id = str(d.get("subscription_id","")).strip()
    if not sub_id:
        return jsonify({"detail": "Subscription ID required"}), 400
    firm = qry("SELECT * FROM firms WHERE sub_id=%s AND is_active=TRUE", (sub_id,), one=True)
    if not firm:
        return jsonify({"detail": "Invalid or inactive Subscription ID"}), 403
    if firm["expires_at"] and datetime.now().date() > firm["expires_at"]:
        return jsonify({"detail": f"Subscription expired on {firm['expires_at']}"}), 403
    # Mark activated
    execute("UPDATE firms SET activated=TRUE WHERE id=%s", (firm["id"],))
    token = jwt.encode({
        "sub_verified": True,
        "firm_id": firm["id"],
        "exp": datetime.utcnow() + timedelta(days=365)
    }, SECRET_KEY, algorithm="HS256")
    return jsonify({
        "subscription_token": token,
        "firm_name": firm["name"],
        "firm_reg_no": firm["reg_no"],
        "activated": True
    })

@app.route("/api/subscription/check", methods=["POST"])
def subscription_check():
    d = request.get_json(silent=True) or {}
    token = str(d.get("subscription_token","")).strip()
    if not token:
        return jsonify({"valid": False})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        firm = qry("SELECT * FROM firms WHERE id=%s AND is_active=TRUE",
                   (payload.get("firm_id"),), one=True)
        if not firm:
            return jsonify({"valid": False})
        if firm["expires_at"] and datetime.now().date() > firm["expires_at"]:
            return jsonify({"valid": False, "reason": "expired"})
        return jsonify({"valid": True, "firm_name": firm["name"]})
    except jwt.PyJWTError:
        return jsonify({"valid": False})

@app.route("/api/firm/identity")
def firm_identity():
    """Frontend calls this to get firm branding after subscription check."""
    token = request.args.get("token","")
    if not token:
        return jsonify({"firm_name":"","firm_reg_no":"","activated":False})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        firm = qry("SELECT * FROM firms WHERE id=%s", (payload.get("firm_id"),), one=True)
        if not firm:
            return jsonify({"firm_name":"","firm_reg_no":"","activated":False})
        return jsonify({
            "firm_name": firm["name"],
            "firm_reg_no": firm["reg_no"] or "",
            "activated": firm["activated"]
        })
    except Exception:
        return jsonify({"firm_name":"","firm_reg_no":"","activated":False})

# ═══════════════════════════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════════════════════════
@app.route("/api/auth/login", methods=["POST"])
def login():
    d = request.get_json()
    username = d.get("username","").strip()
    password = d.get("password","")
    firm_token = d.get("subscription_token","")

    # Resolve firm_id from subscription token or username lookup
    firm_id = None
    if firm_token:
        try:
            p = jwt.decode(firm_token, SECRET_KEY, algorithms=["HS256"])
            firm_id = p.get("firm_id")
        except Exception:
            pass

    if firm_id:
        user = qry("SELECT * FROM users WHERE firm_id=%s AND username=%s",
                   (firm_id, username), one=True)
    else:
        # Multi-firm: find user by username (must be unique across firms or ask for firm)
        user = qry("SELECT * FROM users WHERE username=%s", (username,), one=True)

    if not user or not check_pw(password, user["password_hash"]):
        return jsonify({"detail": "Invalid credentials"}), 401
    if not user["is_active"]:
        return jsonify({"detail": "Account deactivated"}), 403

    # Check firm subscription still valid
    firm = qry("SELECT * FROM firms WHERE id=%s AND is_active=TRUE",
               (user["firm_id"],), one=True)
    if not firm:
        return jsonify({"detail": "Firm account is inactive"}), 403
    if firm["expires_at"] and datetime.now().date() > firm["expires_at"]:
        return jsonify({"detail": f"Firm subscription expired on {firm['expires_at']}"}), 403

    token = make_token(user["id"], user["firm_id"], user["role"])
    log_action(user["firm_id"], user["id"], "LOGIN", ip=request.remote_addr)
    return jsonify({
        "access_token": token,
        "token_type": "bearer",
        "user": _user_out(user),
        "firm": {"name": firm["name"], "reg_no": firm["reg_no"] or ""}
    })

@app.route("/api/auth/me")
@login_required
def auth_me():
    return jsonify(_user_out(g.user))

# ═══════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════
@app.route("/api/dashboard/")
@login_required
def dashboard():
    fid = g.firm_id
    uid = g.user["id"]
    def cnt(sql, params=()):
        return qry(sql, params, one=True)["count"]
    return jsonify({
        "total_clients":      cnt("SELECT COUNT(*) as count FROM clients WHERE firm_id=%s", (fid,)),
        "total_engagements":  cnt("SELECT COUNT(*) as count FROM engagements WHERE firm_id=%s", (fid,)),
        "total_tasks":        cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s", (fid,)),
        "pending_tasks":      cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status='Pending'", (fid,)),
        "in_progress_tasks":  cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status='In Progress'", (fid,)),
        "completed_tasks":    cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status='Completed'", (fid,)),
        "under_review_tasks": cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status='Under Review'", (fid,)),
        "approved_tasks":     cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status='Approved'", (fid,)),
        "rejected_tasks":     cnt("SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status='Rejected'", (fid,)),
        "open_queries":       cnt("SELECT COUNT(*) as count FROM queries WHERE firm_id=%s AND status='Open'", (fid,)),
        "my_pending_tasks":   cnt(
            "SELECT COUNT(*) as count FROM tasks WHERE firm_id=%s AND status IN ('Pending','In Progress') "
            "AND (assignee_id=%s OR id IN (SELECT task_id FROM task_assignees WHERE user_id=%s))",
            (fid, uid, uid)
        ),
    })

# ═══════════════════════════════════════════════════════════════
#  USERS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/users/")
@login_required
def list_users():
    return jsonify([_user_out(u) for u in
                    qry("SELECT * FROM users WHERE firm_id=%s ORDER BY full_name", (g.firm_id,))])

@app.route("/api/users/", methods=["POST"])
@require_role("Admin")
def create_user():
    d = request.get_json()
    fid = g.firm_id
    if qry("SELECT id FROM users WHERE firm_id=%s AND username=%s",
           (fid, d["username"]), one=True):
        return jsonify({"detail": "Username already exists"}), 400
    uid = qry_id(
        "INSERT INTO users (firm_id,username,full_name,email,password_hash,role,client_id) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (fid, d["username"], d["full_name"], d.get("email"),
         hash_pw(d.get("password","audit123")), d.get("role","Member"), d.get("client_id"))
    )
    log_action(fid, g.user["id"], "CREATE_USER", "User", uid,
               f"Created: {d['username']}", request.remote_addr)
    return jsonify(_user_out(qry("SELECT * FROM users WHERE id=%s", (uid,), one=True))), 201

@app.route("/api/users/<int:uid>", methods=["PUT"])
@require_role("Admin")
def update_user(uid):
    d = request.get_json()
    fields, vals = [], []
    for f in ["full_name","email","role","is_active","client_id"]:
        if f in d:
            fields.append(f"{f}=%s")
            vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing to update"}), 400
    vals.append(uid)
    execute(f"UPDATE users SET {','.join(fields)} WHERE id=%s AND firm_id=%s",
            vals + [g.firm_id])
    return jsonify(_user_out(qry("SELECT * FROM users WHERE id=%s", (uid,), one=True)))

@app.route("/api/users/<int:uid>/reset-password", methods=["POST"])
@require_role("Admin")
def reset_password(uid):
    execute("UPDATE users SET password_hash=%s WHERE id=%s AND firm_id=%s",
            (hash_pw("audit123"), uid, g.firm_id))
    return jsonify({"message": "Password reset to 'audit123'"})

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@require_role("Admin")
def delete_user(uid):
    if uid == g.user["id"]:
        return jsonify({"detail": "Cannot delete own account"}), 400
    execute("DELETE FROM users WHERE id=%s AND firm_id=%s", (uid, g.firm_id))
    return jsonify({"message": "User deleted"})

# ═══════════════════════════════════════════════════════════════
#  CLIENTS
# ═══════════════════════════════════════════════════════════════
def _client_out(c):
    return {k: c[k] for k in
            ["id","name","pan","gstin","address","contact_person",
             "contact_phone","contact_email","is_active","created_at"]}

@app.route("/api/clients/")
@login_required
def list_clients():
    return jsonify([_client_out(c) for c in
                    qry("SELECT * FROM clients WHERE firm_id=%s ORDER BY name", (g.firm_id,))])

@app.route("/api/clients/", methods=["POST"])
@login_required
def create_client():
    d = request.get_json()
    cid = qry_id(
        "INSERT INTO clients (firm_id,name,pan,gstin,address,contact_person,"
        "contact_phone,contact_email,created_by_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["name"], d.get("pan"), d.get("gstin"), d.get("address"),
         d.get("contact_person"), d.get("contact_phone"), d.get("contact_email"), g.user["id"])
    )
    log_action(g.firm_id, g.user["id"], "CREATE_CLIENT", "Client", cid,
               f"Created: {d['name']}", request.remote_addr)
    return jsonify(_client_out(qry("SELECT * FROM clients WHERE id=%s", (cid,), one=True))), 201

@app.route("/api/clients/<int:cid>", methods=["PUT"])
@login_required
def update_client(cid):
    d = request.get_json()
    fields, vals = [], []
    for f in ["name","pan","gstin","address","contact_person","contact_phone","contact_email","is_active"]:
        if f in d: fields.append(f"{f}=%s"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing"}), 400
    vals += [cid, g.firm_id]
    execute(f"UPDATE clients SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify(_client_out(qry("SELECT * FROM clients WHERE id=%s", (cid,), one=True)))

@app.route("/api/clients/<int:cid>", methods=["DELETE"])
@require_role("Admin")
def delete_client(cid):
    execute("DELETE FROM clients WHERE id=%s AND firm_id=%s", (cid, g.firm_id))
    return jsonify({"message": "Client deleted"})

# ═══════════════════════════════════════════════════════════════
#  ENGAGEMENTS
# ═══════════════════════════════════════════════════════════════
def _eng_full(e):
    cl = qry("SELECT * FROM clients WHERE id=%s", (e["client_id"],), one=True)
    e["client"] = _client_out(cl) if cl else None
    tl = qry("SELECT * FROM users WHERE id=%s", (e["team_leader_id"],), one=True) if e["team_leader_id"] else None
    e["team_leader"] = _user_out(tl) if tl else None
    return e

@app.route("/api/engagements/")
@login_required
def list_engagements():
    fid = g.firm_id
    if g.user["role"] == "Client":
        cid = g.user.get("client_id")
        if not cid: return jsonify([])
        rows = qry("SELECT * FROM engagements WHERE firm_id=%s AND client_id=%s "
                   "ORDER BY created_at DESC", (fid, cid))
    else:
        rows = qry("SELECT * FROM engagements WHERE firm_id=%s ORDER BY created_at DESC", (fid,))
    return jsonify([_eng_full(e) for e in rows])

@app.route("/api/engagements/<int:eid>")
@login_required
def get_engagement(eid):
    e = qry("SELECT * FROM engagements WHERE id=%s AND firm_id=%s",
            (eid, g.firm_id), one=True)
    if not e: return jsonify({"detail":"Not found"}), 404
    return jsonify(_eng_full(e))

@app.route("/api/engagements/", methods=["POST"])
@login_required
def create_engagement():
    d = request.get_json()
    eid = qry_id(
        "INSERT INTO engagements (firm_id,client_id,title,engagement_type,financial_year,"
        "period_from,period_to,team_leader_id,notes,created_by_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["client_id"], d["title"], d["engagement_type"], d["financial_year"],
         d.get("period_from"), d.get("period_to"), d.get("team_leader_id"),
         d.get("notes"), g.user["id"])
    )
    log_action(g.firm_id, g.user["id"], "CREATE_ENGAGEMENT", "Engagement", eid,
               f"Created: {d['title']}", request.remote_addr)
    return jsonify(_eng_full(qry("SELECT * FROM engagements WHERE id=%s", (eid,), one=True))), 201

@app.route("/api/engagements/<int:eid>", methods=["PUT"])
@login_required
def update_engagement(eid):
    d = request.get_json()
    fields, vals = [], []
    for f in ["title","engagement_type","financial_year","period_from","period_to",
              "team_leader_id","status","notes"]:
        if f in d: fields.append(f"{f}=%s"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing"}), 400
    vals += [eid, g.firm_id]
    execute(f"UPDATE engagements SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify(_eng_full(qry("SELECT * FROM engagements WHERE id=%s", (eid,), one=True)))

@app.route("/api/engagements/<int:eid>", methods=["DELETE"])
@require_role("Admin","Team Leader")
def delete_engagement(eid):
    execute("DELETE FROM engagements WHERE id=%s AND firm_id=%s", (eid, g.firm_id))
    return jsonify({"message": "Deleted"})

@app.route("/api/engagements/<int:eid>/apply-program/<int:pid>", methods=["POST"])
@login_required
def apply_program(eid, pid):
    prog = qry("SELECT * FROM audit_programs WHERE id=%s AND firm_id=%s",
               (pid, g.firm_id), one=True)
    if not prog: return jsonify({"detail":"Program not found"}), 404
    items = qry("SELECT * FROM audit_checklist_items WHERE program_id=%s ORDER BY sr_no", (pid,))
    for item in items:
        tid = qry_id(
            "INSERT INTO tasks (firm_id,engagement_id,checklist_item_id,title,area,priority,"
            "status,working_paper_ref,created_by_id) VALUES (%s,%s,%s,%s,%s,%s,'Pending',%s,%s)",
            (g.firm_id, eid, item["id"], item["description"], item["area"],
             item["priority"], item.get("reference"), g.user["id"])
        )
    return jsonify({"message": f"Created {len(items)} tasks from '{prog['name']}'"})

# ═══════════════════════════════════════════════════════════════
#  TASKS (with multi-assignee)
# ═══════════════════════════════════════════════════════════════
def _task_full(t):
    rows = qry(
        "SELECT u.* FROM task_assignees ta JOIN users u ON u.id=ta.user_id "
        "WHERE ta.task_id=%s ORDER BY u.full_name", (t["id"],))
    t["assignees"] = [_user_out(u) for u in rows]
    t["assignee_ids"] = [u["id"] for u in rows]
    return t

def _sync_assignees(tid, user_ids):
    execute("DELETE FROM task_assignees WHERE task_id=%s", (tid,), commit=False)
    for uid in user_ids:
        try:
            execute("INSERT INTO task_assignees (task_id,user_id) VALUES (%s,%s)",
                    (tid, uid), commit=False)
        except Exception: pass
    primary = user_ids[0] if user_ids else None
    execute("UPDATE tasks SET assignee_id=%s WHERE id=%s", (primary, tid), commit=False)
    get_db().commit()

@app.route("/api/tasks/")
@login_required
def list_tasks():
    sql = "SELECT * FROM tasks WHERE firm_id=%s"
    params = [g.firm_id]
    if request.args.get("engagement_id"):
        sql += " AND engagement_id=%s"; params.append(request.args["engagement_id"])
    if request.args.get("status"):
        sql += " AND status=%s"; params.append(request.args["status"])
    sql += " ORDER BY created_at DESC"
    return jsonify([_task_full(t) for t in qry(sql, params)])

@app.route("/api/tasks/<int:tid>")
@login_required
def get_task(tid):
    t = qry("SELECT * FROM tasks WHERE id=%s AND firm_id=%s", (tid, g.firm_id), one=True)
    if not t: return jsonify({"detail":"Not found"}), 404
    return jsonify(_task_full(t))

@app.route("/api/tasks/", methods=["POST"])
@login_required
def create_task():
    d = request.get_json()
    ids = d.get("assignee_ids") or ([d["assignee_id"]] if d.get("assignee_id") else [])
    primary = ids[0] if ids else None
    tid = qry_id(
        "INSERT INTO tasks (firm_id,engagement_id,title,description,area,assignee_id,"
        "priority,due_date,working_paper_ref,created_by_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["engagement_id"], d["title"], d.get("description"), d.get("area"),
         primary, d.get("priority","Medium"), d.get("due_date"), d.get("working_paper_ref"), g.user["id"])
    )
    _sync_assignees(tid, ids)
    log_action(g.firm_id, g.user["id"], "CREATE_TASK", "Task", tid,
               f"Created: {d['title']}", request.remote_addr)
    return jsonify(_task_full(qry("SELECT * FROM tasks WHERE id=%s", (tid,), one=True))), 201

@app.route("/api/tasks/<int:tid>", methods=["PUT"])
@login_required
def update_task(tid):
    d = request.get_json()
    fields, vals = [], []
    if "assignee_ids" in d:
        _sync_assignees(tid, d["assignee_ids"] or [])
        fields.append("assignee_id=%s")
        vals.append(d["assignee_ids"][0] if d["assignee_ids"] else None)
    for f in ["title","description","area","status","priority","due_date","working_paper_ref"]:
        if f in d: fields.append(f"{f}=%s"); vals.append(d[f])
    if d.get("status") == "Completed":
        fields.append("completed_at=%s"); vals.append(datetime.utcnow())
    if fields:
        vals += [tid, g.firm_id]
        execute(f"UPDATE tasks SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify(_task_full(qry("SELECT * FROM tasks WHERE id=%s", (tid,), one=True)))

@app.route("/api/tasks/<int:tid>", methods=["DELETE"])
@login_required
def delete_task(tid):
    execute("DELETE FROM tasks WHERE id=%s AND firm_id=%s", (tid, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  COMMENTS
# ═══════════════════════════════════════════════════════════════
def _comment_full(c):
    a = qry("SELECT * FROM users WHERE id=%s", (c["author_id"],), one=True)
    c["author"] = _user_out(a) if a else None
    return c

@app.route("/api/comments/")
@login_required
def list_comments():
    tid = request.args.get("task_id")
    rows = qry("SELECT * FROM comments WHERE task_id=%s AND firm_id=%s ORDER BY created_at",
               (tid, g.firm_id))
    return jsonify([_comment_full(c) for c in rows])

@app.route("/api/comments/", methods=["POST"])
@login_required
def create_comment():
    d = request.get_json()
    is_q = bool(d.get("is_query"))
    cid = qry_id(
        "INSERT INTO comments (firm_id,task_id,author_id,content,is_query) VALUES (%s,%s,%s,%s,%s)",
        (g.firm_id, d["task_id"], g.user["id"], d["content"], is_q)
    )
    if is_q:
        task = qry("SELECT * FROM tasks WHERE id=%s", (d["task_id"],), one=True)
        if task:
            max_sr = qry("SELECT COALESCE(MAX(sr_no),0) as m FROM queries "
                         "WHERE engagement_id=%s", (task["engagement_id"],), one=True)["m"]
            execute(
                "INSERT INTO queries (firm_id,engagement_id,comment_id,sr_no,query_text,"
                "raised_by_id,task_reference) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (g.firm_id, task["engagement_id"], cid, max_sr+1, d["content"],
                 g.user["id"], f"Task #{task['id']}: {task['title']}")
            )
    return jsonify(_comment_full(qry("SELECT * FROM comments WHERE id=%s", (cid,), one=True))), 201

@app.route("/api/comments/<int:cid>", methods=["PUT"])
@login_required
def update_comment(cid):
    d = request.get_json()
    comment = qry("SELECT * FROM comments WHERE id=%s AND firm_id=%s", (cid, g.firm_id), one=True)
    if not comment: return jsonify({"detail":"Not found"}), 404
    if comment["author_id"] != g.user["id"]:
        return jsonify({"detail":"Can only edit own comments"}), 403
    execute("UPDATE comments SET content=%s, is_query=%s, updated_at=%s WHERE id=%s",
            (d.get("content", comment["content"]), bool(d.get("is_query")), datetime.utcnow(), cid))
    return jsonify(_comment_full(qry("SELECT * FROM comments WHERE id=%s", (cid,), one=True)))

@app.route("/api/comments/<int:cid>", methods=["DELETE"])
@login_required
def delete_comment(cid):
    comment = qry("SELECT * FROM comments WHERE id=%s AND firm_id=%s", (cid, g.firm_id), one=True)
    if not comment: return jsonify({"detail":"Not found"}), 404
    if comment["author_id"] != g.user["id"]:
        return jsonify({"detail":"Can only delete own comments"}), 403
    execute("DELETE FROM queries WHERE comment_id=%s", (cid,))
    execute("DELETE FROM comments WHERE id=%s", (cid,))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  QUERIES
# ═══════════════════════════════════════════════════════════════
def _query_full(q):
    rb = qry("SELECT * FROM users WHERE id=%s", (q["raised_by_id"],), one=True)
    q["raised_by"] = _user_out(rb) if rb else None
    return q

@app.route("/api/queries/")
@login_required
def list_queries():
    sql = "SELECT * FROM queries WHERE firm_id=%s"
    params = [g.firm_id]
    if request.args.get("engagement_id"):
        sql += " AND engagement_id=%s"; params.append(request.args["engagement_id"])
    if request.args.get("status"):
        sql += " AND status=%s"; params.append(request.args["status"])
    sql += " ORDER BY sr_no"
    return jsonify([_query_full(q) for q in qry(sql, params)])

@app.route("/api/queries/", methods=["POST"])
@login_required
def create_query():
    d = request.get_json()
    max_sr = qry("SELECT COALESCE(MAX(sr_no),0) as m FROM queries WHERE engagement_id=%s",
                 (d["engagement_id"],), one=True)["m"]
    qid = qry_id(
        "INSERT INTO queries (firm_id,engagement_id,sr_no,query_text,raised_by_id,task_reference) "
        "VALUES (%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["engagement_id"], max_sr+1, d["query_text"], g.user["id"], d.get("task_reference"))
    )
    return jsonify(_query_full(qry("SELECT * FROM queries WHERE id=%s", (qid,), one=True))), 201

@app.route("/api/queries/<int:qid>", methods=["PUT"])
@login_required
def update_query(qid):
    d = request.get_json()
    fields, vals = [], []
    if "response" in d:
        fields += ["response=%s","responded_by_id=%s","responded_date=%s"]
        vals += [d["response"], g.user["id"], datetime.utcnow()]
        if "status" not in d: fields.append("status='Responded'")
    if "status" in d: fields.append("status=%s"); vals.append(d["status"])
    vals += [qid, g.firm_id]
    execute(f"UPDATE queries SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify(_query_full(qry("SELECT * FROM queries WHERE id=%s", (qid,), one=True)))

@app.route("/api/queries/<int:qid>", methods=["DELETE"])
@require_role("Admin","Team Leader")
def delete_query(qid):
    execute("DELETE FROM queries WHERE id=%s AND firm_id=%s", (qid, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  QUERY SHEET EXCEL EXPORT
#  Place this immediately AFTER the /api/queries/<qid> DELETE route
# ═══════════════════════════════════════════════════════════════

@app.route("/api/queries/export/<int:eid>")
@login_required
def export_queries(eid):
    import io, tempfile, os
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"detail": "openpyxl not installed"}), 500

    eng = qry("""SELECT e.*, c.name as client_name
                 FROM engagements e JOIN clients c ON e.client_id=c.id
                 WHERE e.id=%s AND e.firm_id=%s""",
              (eid, g.firm_id), one=True)
    if not eng:
        return jsonify({"detail": "Engagement not found"}), 404

    queries = qry("""SELECT q.*, u.full_name as raised_by_name
                     FROM queries q LEFT JOIN users u ON q.raised_by_id=u.id
                     WHERE q.engagement_id=%s AND q.firm_id=%s
                     ORDER BY q.sr_no""", (eid, g.firm_id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Query Sheet"

    navy    = "003366"
    gold    = "DAA520"
    light   = "EBF4FF"
    thin    = Side(style="thin", color="D1DAEA")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1,
                value=f"Query Sheet — {eng.get('client_name','')} | {eng.get('title','')} | FY {eng.get('financial_year','')}")
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=navy)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header row ─────────────────────────────────────────────
    headers = ["Sr. No.", "Query", "Response", "Status", "Raised By", "Date", "Task Reference"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1a4d80")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr
    ws.row_dimensions[2].height = 18

    # ── Data rows ──────────────────────────────────────────────
    for i, q in enumerate(queries):
        row = i + 3
        vals = [
            q.get("sr_no", ""),
            q.get("query_text", "") or "",
            q.get("response", "") or "Pending",
            q.get("status", "") or "",
            q.get("raised_by_name", "") or "",
            str(q.get("raised_date", "") or "")[:10],
            q.get("task_reference", "") or "",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=str(val) if val is not None else "")
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=light if i % 2 == 0 else "FFFFFF")
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border    = bdr
        ws.row_dimensions[row].height = 40

    # ── Column widths ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30

    # ── Save to temp file and serve ────────────────────────────
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False,
        dir=os.path.join(os.path.dirname(__file__), "static"))
    tmp.close()
    wb.save(tmp.name)

    client_name = (eng.get("client_name") or "Client").replace(" ", "_")
    fy = (eng.get("financial_year") or "").replace("-", "_")
    download_name = f"Query_Sheet_{client_name}_{fy}.xlsx"

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
  
# ═══════════════════════════════════════════════════════════════
#  AUDIT PROGRAMS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/programs/")
@login_required
def list_programs():
    progs = qry("SELECT * FROM audit_programs WHERE firm_id=%s AND is_active=TRUE", (g.firm_id,))
    for p in progs:
        p["checklist_items"] = qry(
            "SELECT * FROM audit_checklist_items WHERE program_id=%s ORDER BY sr_no", (p["id"],))
    return jsonify(progs)

@app.route("/api/programs/", methods=["POST"])
@require_role("Admin","Team Leader")
def create_program():
    d = request.get_json()
    pid = qry_id(
        "INSERT INTO audit_programs (firm_id,name,engagement_type,description) VALUES (%s,%s,%s,%s)",
        (g.firm_id, d["name"], d["engagement_type"], d.get("description"))
    )
    for item in d.get("checklist_items",[]):
        execute(
            "INSERT INTO audit_checklist_items (program_id,sr_no,area,description,reference,priority) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (pid, item["sr_no"], item["area"], item["description"],
             item.get("reference"), item.get("priority","Medium"))
        )
    return jsonify({"id": pid, "name": d["name"]}), 201

@app.route("/api/programs/<int:pid>", methods=["DELETE"])
@require_role("Admin","Team Leader")
def delete_program(pid):
    execute("DELETE FROM audit_checklist_items WHERE program_id=%s", (pid,))
    execute("DELETE FROM audit_programs WHERE id=%s AND firm_id=%s", (pid, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  REVIEWS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/reviews/")
@login_required
def list_reviews():
    tid = request.args.get("task_id")
    rows = qry(
        "SELECT r.*, u.full_name, u.username FROM reviews r "
        "JOIN users u ON r.reviewer_id=u.id WHERE r.task_id=%s AND r.firm_id=%s "
        "ORDER BY r.reviewed_at DESC", (tid, g.firm_id))
    return jsonify(rows)

@app.route("/api/reviews/", methods=["POST"])
@require_role("Admin","Team Leader")
def submit_review():
    d = request.get_json()
    task = qry("SELECT * FROM tasks WHERE id=%s AND firm_id=%s", (d["task_id"], g.firm_id), one=True)
    if not task: return jsonify({"detail":"Task not found"}), 404
    rid = qry_id(
        "INSERT INTO reviews (firm_id,task_id,reviewer_id,action,remarks) VALUES (%s,%s,%s,%s,%s)",
        (g.firm_id, d["task_id"], g.user["id"], d["action"], d.get("remarks"))
    )
    new_status = "Approved" if d["action"] == "Approved" else "Rejected"
    execute("UPDATE tasks SET status=%s WHERE id=%s", (new_status, d["task_id"]))
    log_action(g.firm_id, g.user["id"], "REVIEW_TASK", "Review", rid,
               f"Task #{d['task_id']}: {d['action']}", request.remote_addr)
    return jsonify({"id": rid, "action": d["action"]}), 201

# ═══════════════════════════════════════════════════════════════
#  TIME LOGS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/time-logs/")
@login_required
def list_time_logs():
    sql = "SELECT tl.*, u.full_name FROM time_logs tl JOIN users u ON u.id=tl.user_id WHERE tl.firm_id=%s"
    params = [g.firm_id]
    if request.args.get("task_id"):
        sql += " AND tl.task_id=%s"; params.append(request.args["task_id"])
    sql += " ORDER BY tl.date DESC"
    return jsonify(qry(sql, params))

@app.route("/api/time-logs/", methods=["POST"])
@login_required
def create_time_log():
    d = request.get_json()
    hours = float(d.get("hours", 0))
    if hours <= 0: return jsonify({"detail":"Hours > 0 required"}), 400
    lid = qry_id(
        "INSERT INTO time_logs (firm_id,task_id,user_id,date,hours,note) VALUES (%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["task_id"], g.user["id"],
         d.get("date", datetime.utcnow().strftime("%Y-%m-%d")), hours, d.get("note"))
    )
    return jsonify({"id": lid, "hours": hours}), 201

@app.route("/api/time-logs/<int:lid>", methods=["DELETE"])
@login_required
def delete_time_log(lid):
    execute("DELETE FROM time_logs WHERE id=%s AND firm_id=%s", (lid, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  INVOICES
# ═══════════════════════════════════════════════════════════════
def _inv_full(inv):
    cl = qry("SELECT * FROM clients WHERE id=%s", (inv["client_id"],), one=True)
    inv["client"] = _client_out(cl) if cl else None
    eng = qry("SELECT * FROM engagements WHERE id=%s", (inv["engagement_id"],), one=True)
    inv["engagement_title"] = eng["title"] if eng else ""
    inv["engagement_fy"]    = eng["financial_year"] if eng else ""
    return inv

@app.route("/api/invoices/")
@login_required
def list_invoices():
    sql = "SELECT * FROM invoices WHERE firm_id=%s"
    params = [g.firm_id]
    if request.args.get("payment_status"):
        sql += " AND payment_status=%s"; params.append(request.args["payment_status"])
    sql += " ORDER BY invoice_date DESC"
    return jsonify([_inv_full(i) for i in qry(sql, params)])

@app.route("/api/invoices/summary/")
@login_required
def invoice_summary():
    row = qry(
        "SELECT COALESCE(SUM(total_amount),0) as total_billed, "
        "COALESCE(SUM(CASE WHEN payment_status='Paid' THEN total_amount ELSE 0 END),0) as total_received, "
        "COALESCE(SUM(CASE WHEN payment_status='Unpaid' THEN total_amount ELSE 0 END),0) as total_outstanding, "
        "COUNT(*) as total_invoices FROM invoices WHERE firm_id=%s", (g.firm_id,), one=True)
    return jsonify(row)

@app.route("/api/invoices/", methods=["POST"])
@require_role("Admin","Team Leader")
def create_invoice():
    d = request.get_json()
    amount  = float(d.get("amount",0))
    gst_pct = float(d.get("gst_percent",18))
    gst_amt = round(amount * gst_pct / 100, 2)
    total   = round(amount + gst_amt, 2)
    max_id  = qry("SELECT COALESCE(MAX(id),0) as m FROM invoices WHERE firm_id=%s",
                  (g.firm_id,), one=True)["m"]
    inv_no  = d.get("invoice_no") or f"INV/{datetime.now().strftime('%y%m')}/{max_id+1:04d}"
    iid = qry_id(
        "INSERT INTO invoices (firm_id,engagement_id,client_id,invoice_no,invoice_date,"
        "description,amount,gst_percent,gst_amount,total_amount,payment_status,created_by_id) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["engagement_id"], d["client_id"], inv_no, d["invoice_date"],
         d.get("description"), amount, gst_pct, gst_amt, total,
         d.get("payment_status","Unpaid"), g.user["id"])
    )
    return jsonify(_inv_full(qry("SELECT * FROM invoices WHERE id=%s", (iid,), one=True))), 201

@app.route("/api/invoices/<int:iid>", methods=["PUT"])
@require_role("Admin","Team Leader")
def update_invoice(iid):
    d = request.get_json()
    fields, vals = [], []
    for f in ["invoice_date","description","payment_status","payment_date","payment_note"]:
        if f in d: fields.append(f"{f}=%s"); vals.append(d[f])
    if "amount" in d or "gst_percent" in d:
        inv = qry("SELECT * FROM invoices WHERE id=%s", (iid,), one=True)
        amt = float(d.get("amount", inv["amount"]))
        gp  = float(d.get("gst_percent", inv["gst_percent"]))
        ga  = round(amt*gp/100, 2)
        tot = round(amt+ga, 2)
        fields += ["amount=%s","gst_percent=%s","gst_amount=%s","total_amount=%s"]
        vals += [amt, gp, ga, tot]
    if not fields: return jsonify({"detail":"Nothing"}), 400
    vals += [iid, g.firm_id]
    execute(f"UPDATE invoices SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify(_inv_full(qry("SELECT * FROM invoices WHERE id=%s", (iid,), one=True)))

@app.route("/api/invoices/<int:iid>", methods=["DELETE"])
@require_role("Admin")
def delete_invoice(iid):
    execute("DELETE FROM invoices WHERE id=%s AND firm_id=%s", (iid, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  DOCUMENT REGISTER
# ═══════════════════════════════════════════════════════════════
@app.route("/api/doc-register/")
@login_required
def list_doc_register():
    sql = "SELECT dr.*, c.name as client_name, e.title as engagement_title FROM doc_register dr " \
          "LEFT JOIN clients c ON dr.client_id=c.id LEFT JOIN engagements e ON dr.engagement_id=e.id " \
          "WHERE dr.firm_id=%s ORDER BY dr.doc_date DESC"
    return jsonify(qry(sql, (g.firm_id,)))

@app.route("/api/doc-register/", methods=["POST"])
@login_required
def create_doc_entry():
    d = request.get_json()
    did = qry_id(
        "INSERT INTO doc_register (firm_id,engagement_id,client_id,doc_type,doc_name,"
        "doc_category,doc_date,received_from,sent_to,reference_no,status,remarks,created_by_id) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d.get("engagement_id"), d.get("client_id"), d["doc_type"], d["doc_name"],
         d.get("doc_category"), d["doc_date"], d.get("received_from"), d.get("sent_to"),
         d.get("reference_no"), d.get("status","Received"), d.get("remarks"), g.user["id"])
    )
    return jsonify({"id": did}), 201

@app.route("/api/doc-register/<int:did>", methods=["PUT"])
@login_required
def update_doc_entry(did):
    d = request.get_json()
    fields, vals = [], []
    for f in ["doc_name","doc_category","doc_date","received_from","sent_to",
              "reference_no","status","remarks"]:
        if f in d: fields.append(f"{f}=%s"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing"}), 400
    vals += [did, g.firm_id]
    execute(f"UPDATE doc_register SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify({"message":"Updated"})

@app.route("/api/doc-register/<int:did>", methods=["DELETE"])
@login_required
def delete_doc_entry(did):
    execute("DELETE FROM doc_register WHERE id=%s AND firm_id=%s", (did, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  COMPLIANCE CALENDAR
# ═══════════════════════════════════════════════════════════════
@app.route("/api/compliance-calendar/")
@login_required
def list_compliance():
    sql = "SELECT * FROM compliance_calendar WHERE firm_id=%s"
    params = [g.firm_id]
    if request.args.get("category"):
        sql += " AND category=%s"; params.append(request.args["category"])
    if request.args.get("status"):
        sql += " AND status=%s"; params.append(request.args["status"])
    sql += " ORDER BY due_date"
    return jsonify(qry(sql, params))

@app.route("/api/compliance-calendar/", methods=["POST"])
@require_role("Admin","Team Leader")
def create_compliance():
    d = request.get_json()
    cid = qry_id(
        "INSERT INTO compliance_calendar (firm_id,title,category,due_date,description,"
        "financial_year,engagement_id,is_recurring,status,created_by_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, d["title"], d["category"], d["due_date"], d.get("description"),
         d.get("financial_year"), d.get("engagement_id"), bool(d.get("is_recurring")),
         d.get("status","Upcoming"), g.user["id"])
    )
    return jsonify({"id": cid}), 201

@app.route("/api/compliance-calendar/<int:cid>", methods=["PUT"])
@require_role("Admin","Team Leader")
def update_compliance(cid):
    d = request.get_json()
    fields, vals = [], []
    for f in ["title","category","due_date","description","financial_year","status","is_recurring"]:
        if f in d: fields.append(f"{f}=%s"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing"}), 400
    vals += [cid, g.firm_id]
    execute(f"UPDATE compliance_calendar SET {','.join(fields)} WHERE id=%s AND firm_id=%s", vals)
    return jsonify({"message":"Updated"})

@app.route("/api/compliance-calendar/<int:cid>", methods=["DELETE"])
@require_role("Admin","Team Leader")
def delete_compliance(cid):
    execute("DELETE FROM compliance_calendar WHERE id=%s AND firm_id=%s", (cid, g.firm_id))
    return jsonify({"message":"Deleted"})

@app.route("/api/compliance-calendar/seed-fy", methods=["POST"])
@require_role("Admin","Team Leader")
def seed_compliance_fy():
    d = request.get_json()
    fy = d.get("financial_year","")
    if not fy: return jsonify({"detail":"financial_year required"}), 400
    try:
        start_yr = int(fy.split("-")[0]); end_yr = start_yr + 1
    except Exception:
        return jsonify({"detail":"Use YYYY-YY format"}), 400
    deadlines = [
        ("GSTR-1 (Quarterly) Q1","GST",f"{end_yr-1}-07-31"),
        ("GSTR-1 (Quarterly) Q2","GST",f"{end_yr-1}-10-31"),
        ("GSTR-1 (Quarterly) Q3","GST",f"{end_yr}-01-31"),
        ("GSTR-1 (Quarterly) Q4","GST",f"{end_yr}-04-30"),
        ("GSTR-9 Annual Return","GST",f"{end_yr}-12-31"),
        ("TDS Q1 Return","TDS",f"{end_yr-1}-07-31"),
        ("TDS Q2 Return","TDS",f"{end_yr-1}-10-31"),
        ("TDS Q3 Return","TDS",f"{end_yr}-01-31"),
        ("TDS Q4 Return","TDS",f"{end_yr}-05-31"),
        ("TDS Certificate Form 16","TDS",f"{end_yr}-06-15"),
        ("ITR Filing — Individual/HUF","ITR",f"{end_yr}-07-31"),
        ("ITR Filing — Companies/Tax Audit","ITR",f"{end_yr}-10-31"),
        ("Advance Tax Q1 (15%)","ITR",f"{end_yr-1}-06-15"),
        ("Advance Tax Q2 (45%)","ITR",f"{end_yr-1}-09-15"),
        ("Advance Tax Q3 (75%)","ITR",f"{end_yr-1}-12-15"),
        ("Advance Tax Q4 (100%)","ITR",f"{end_yr}-03-15"),
        ("ROC Annual Return MGT-7","ROC",f"{end_yr}-11-29"),
        ("ROC Financial Statements AOC-4","ROC",f"{end_yr}-10-29"),
        ("AGM Holding Deadline","ROC",f"{end_yr}-09-30"),
    ]
    added = 0
    for title, cat, due in deadlines:
        exists = qry("SELECT id FROM compliance_calendar WHERE firm_id=%s AND title=%s AND financial_year=%s",
                     (g.firm_id, title, fy), one=True)
        if not exists:
            execute(
                "INSERT INTO compliance_calendar (firm_id,title,category,due_date,financial_year,"
                "is_recurring,status,created_by_id) VALUES (%s,%s,%s,%s,%s,TRUE,'Upcoming',%s)",
                (g.firm_id, title, cat, due, fy, g.user["id"])
            )
            added += 1
    return jsonify({"message": f"Seeded {added} dates for FY {fy}"})

# ═══════════════════════════════════════════════════════════════
#  ALERTS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/alerts/")
@login_required
def get_alerts():
    fid = g.firm_id; uid = g.user["id"]
    alerts = []
    today = datetime.now().date()
    overdue = qry(
        "SELECT t.id, t.title, t.due_date FROM tasks t "
        "WHERE t.firm_id=%s AND t.due_date < %s "
        "AND t.status NOT IN ('Completed','Approved') LIMIT 10",
        (fid, today))
    for t in overdue:
        alerts.append({"type":"overdue_task","severity":"high",
                       "message":f"OVERDUE: {t['title']} (due {t['due_date']})","task_id":t["id"]})
    due_soon = qry(
        "SELECT t.id, t.title, t.due_date FROM tasks t "
        "WHERE t.firm_id=%s AND t.due_date BETWEEN %s AND %s "
        "AND t.status NOT IN ('Completed','Approved') LIMIT 10",
        (fid, today, today + timedelta(days=3)))
    for t in due_soon:
        alerts.append({"type":"due_soon","severity":"warn",
                       "message":f"Due soon: {t['title']} on {t['due_date']}","task_id":t["id"]})
    comp = qry(
        "SELECT * FROM compliance_calendar WHERE firm_id=%s AND status='Upcoming' "
        "AND due_date BETWEEN %s AND %s ORDER BY due_date LIMIT 5",
        (fid, today, today + timedelta(days=7)))
    for c in comp:
        alerts.append({"type":"compliance","severity":"info",
                       "message":f"Compliance due: {c['title']} on {c['due_date']}"})
    return jsonify(alerts)

# ═══════════════════════════════════════════════════════════════
#  AUDIT LOGS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/audit-logs/")
@require_role("Admin","Team Leader")
def list_audit_logs():
    limit = request.args.get("limit", 200, type=int)
    rows = qry(
        "SELECT al.*, u.full_name FROM audit_logs al "
        "LEFT JOIN users u ON al.user_id=u.id "
        "WHERE al.firm_id=%s ORDER BY al.timestamp DESC LIMIT %s",
        (g.firm_id, limit))
    return jsonify(rows)

@app.route("/api/audit-logs/clear", methods=["DELETE"])
@require_role("Admin")
def clear_audit_logs():
    execute("DELETE FROM audit_logs WHERE firm_id=%s", (g.firm_id,))
    return jsonify({"message":"Logs cleared"})

# ═══════════════════════════════════════════════════════════════
#  REPORTS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/reports/overview")
@login_required
def report_overview():
    fid = g.firm_id
    today = datetime.now().date()
    eng_stats = qry("""
        SELECT e.id, e.title, c.name as client_name, e.financial_year,
               COUNT(t.id) as total_tasks,
               SUM(CASE WHEN t.status IN ('Completed','Approved') THEN 1 ELSE 0 END) as done_tasks,
               SUM(CASE WHEN t.due_date < CURRENT_DATE AND t.status NOT IN ('Completed','Approved') THEN 1 ELSE 0 END) as overdue_tasks,
               COALESCE(SUM(tl.hours),0) as total_hours
        FROM engagements e JOIN clients c ON e.client_id=c.id
        LEFT JOIN tasks t ON t.engagement_id=e.id
        LEFT JOIN time_logs tl ON tl.task_id=t.id
        WHERE e.firm_id=%s GROUP BY e.id, c.name ORDER BY e.created_at DESC
    """, (fid,))
    staff = qry("""
        SELECT u.id, u.full_name, u.role,
               COUNT(DISTINCT t.id) as assigned_tasks,
               SUM(CASE WHEN t.status IN ('Completed','Approved') THEN 1 ELSE 0 END) as completed_tasks,
               COALESCE(SUM(tl.hours),0) as logged_hours
        FROM users u LEFT JOIN tasks t ON t.assignee_id=u.id
        LEFT JOIN time_logs tl ON tl.user_id=u.id
        WHERE u.firm_id=%s AND u.role IN ('Admin','Team Leader','Member')
        GROUP BY u.id ORDER BY completed_tasks DESC
    """, (fid,))
    inv = qry(
        "SELECT COALESCE(SUM(total_amount),0) as billed, "
        "COALESCE(SUM(CASE WHEN payment_status='Paid' THEN total_amount ELSE 0 END),0) as received, "
        "COALESCE(SUM(CASE WHEN payment_status='Unpaid' THEN total_amount ELSE 0 END),0) as outstanding "
        "FROM invoices WHERE firm_id=%s", (fid,), one=True)
    upcoming = qry(
        "SELECT * FROM compliance_calendar WHERE firm_id=%s AND status='Upcoming' "
        "AND due_date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days' "
        "ORDER BY due_date LIMIT 10", (fid,))
    due_soon = qry(
        "SELECT t.id, t.title, t.due_date, t.priority, u.full_name as assignee_name, "
        "e.title as engagement_title FROM tasks t "
        "LEFT JOIN users u ON t.assignee_id=u.id "
        "LEFT JOIN engagements e ON t.engagement_id=e.id "
        "WHERE t.firm_id=%s AND t.due_date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '7 days' "
        "AND t.status NOT IN ('Completed','Approved') ORDER BY t.due_date LIMIT 20", (fid,))
    return jsonify({
        "engagement_stats": eng_stats,
        "staff_performance": staff,
        "invoice_summary": inv,
        "upcoming_compliance": upcoming,
        "tasks_due_soon": due_soon,
        "query_avg_tat_days": None
    })
# ═══════════════════════════════════════════════════════════════
#  AUDIT BOOKLET GENERATOR
# ═══════════════════════════════════════════════════════════════

@app.route("/api/booklet/generate/<int:eid>")
@login_required
def gen_booklet(eid):
    import tempfile, os
    try:
        from booklet_generator import generate_booklet
    except ImportError:
        return jsonify({"detail": "python-docx not installed or booklet_generator.py missing"}), 500

    # Engagement + client details
    eng = qry("""SELECT e.*, c.name as client_name, c.pan, c.gstin
                 FROM engagements e JOIN clients c ON e.client_id=c.id
                 WHERE e.id=%s AND e.firm_id=%s""",
              (eid, g.firm_id), one=True)
    if not eng:
        return jsonify({"detail": "Not found"}), 404

    # Build team list
    seen_ids = set()
    team = []
    if eng.get("team_leader_id"):
        tl = qry("SELECT * FROM users WHERE id=%s", (eng["team_leader_id"],), one=True)
        if tl:
            team.append({"full_name": tl["full_name"], "role": "Team Leader",
                         "email": tl.get("email", "")})
            seen_ids.add(tl["id"])

    for m in qry("""SELECT u.* FROM engagement_teams et
                    JOIN users u ON et.user_id=u.id
                    WHERE et.engagement_id=%s""", (eid,)):
        if m["id"] not in seen_ids:
            team.append({"full_name": m["full_name"], "role": m.get("role", "Member"),
                         "email": m.get("email", "")})
            seen_ids.add(m["id"])

    for m in qry("""SELECT DISTINCT u.* FROM task_assignees ta
                    JOIN tasks t ON ta.task_id=t.id
                    JOIN users u ON ta.user_id=u.id
                    WHERE t.engagement_id=%s""", (eid,)):
        if m["id"] not in seen_ids:
            team.append({"full_name": m["full_name"], "role": m.get("role", "Audit Member"),
                         "email": m.get("email", "")})
            seen_ids.add(m["id"])

    # Tasks
    tasks_raw = qry("SELECT * FROM tasks WHERE engagement_id=%s AND firm_id=%s ORDER BY id",
                    (eid, g.firm_id))

    # Assignee names per task
    task_assignee_names = {}
    if tasks_raw:
        tids = [t["id"] for t in tasks_raw]
        ph = ",".join(["%s"] * len(tids))
        for row in qry(f"""SELECT ta.task_id, u.full_name
                           FROM task_assignees ta JOIN users u ON ta.user_id=u.id
                           WHERE ta.task_id IN ({ph})
                           ORDER BY u.full_name""", tids):
            task_assignee_names.setdefault(row["task_id"], []).append(row["full_name"])
        for t in tasks_raw:
            if t["id"] not in task_assignee_names and t.get("assignee_id"):
                u = qry("SELECT full_name FROM users WHERE id=%s",
                        (t["assignee_id"],), one=True)
                if u:
                    task_assignee_names[t["id"]] = [u["full_name"]]

    tasks = [{
        "id":                t["id"],
        "title":             t["title"],
        "area":              t.get("area", ""),
        "status":            t["status"],
        "priority":          t.get("priority", ""),
        "due_date":          str(t["due_date"]) if t.get("due_date") else "",
        "working_paper_ref": t.get("working_paper_ref", ""),
        "description":       t.get("description", ""),
        "assignee_name":     ", ".join(task_assignee_names.get(t["id"], [])) or "Unassigned"
    } for t in tasks_raw]

    tids_all = [t["id"] for t in tasks_raw]

    # Comments by task
    cbt = {}
    if tids_all:
        ph = ",".join(["%s"] * len(tids_all))
        for c in qry(f"""SELECT c.*, u.full_name as author_name
                         FROM comments c JOIN users u ON c.author_id=u.id
                         WHERE c.task_id IN ({ph})
                         ORDER BY c.created_at""", tids_all):
            cbt.setdefault(c["task_id"], []).append({
                "author_name": c["author_name"],
                "content":     c["content"],
                "is_query":    c["is_query"],
                "created_at":  str(c["created_at"])[:16] if c.get("created_at") else ""
            })

    # Reviews by task
    rbt = {}
    if tids_all:
        ph = ",".join(["%s"] * len(tids_all))
        for r in qry(f"""SELECT r.*, u.full_name as reviewer_name
                         FROM reviews r JOIN users u ON r.reviewer_id=u.id
                         WHERE r.task_id IN ({ph})
                         ORDER BY r.reviewed_at""", tids_all):
            rbt.setdefault(r["task_id"], []).append({
                "reviewer_name": r["reviewer_name"],
                "action":        r["action"],
                "remarks":       r.get("remarks", ""),
                "reviewed_at":   str(r["reviewed_at"])[:16] if r.get("reviewed_at") else ""
            })

    # Queries
    queries = qry("""SELECT q.*, u.full_name as raised_by_name
                     FROM queries q LEFT JOIN users u ON q.raised_by_id=u.id
                     WHERE q.engagement_id=%s AND q.firm_id=%s
                     ORDER BY q.sr_no""", (eid, g.firm_id))

    # Audit program checklist items — Planning first, Completion last, others by sr_no
    audit_programs_raw = qry("""
        SELECT ap.sr_no, ap.area, ap.description, ap.reference, ap.priority,
               ap.program_id
        FROM audit_checklist_items ap
        JOIN audit_programs prog ON ap.program_id = prog.id
        WHERE prog.firm_id = %s
        ORDER BY
            CASE
                WHEN LOWER(ap.area) LIKE 'planning%%'   THEN 1
                WHEN LOWER(ap.area) LIKE 'completion%%' THEN 3
                ELSE 2
            END,
            ap.program_id,
            ap.sr_no
    """, (g.firm_id,))

    audit_programs_list = [{
        "sr_no":       row["sr_no"],
        "area":        row["area"],
        "description": row["description"],
        "reference":   row.get("reference", "") or "",
        "priority":    row.get("priority", "") or "",
        "program_id":  row["program_id"],
    } for row in audit_programs_raw]

    # Get firm details
    firm = qry("SELECT * FROM firms WHERE id=%s", (g.firm_id,), one=True)
    firm_name_str = firm["name"] if firm else ""
    firm_reg_no   = firm.get("reg_no", "") if firm else ""

    # Build queries list once
    queries_list = [{
        "sr_no":          q["sr_no"],
        "query_text":     q["query_text"],
        "response":       q.get("response", ""),
        "status":         q["status"],
        "raised_by_name": q.get("raised_by_name", ""),
        "raised_date":    str(q.get("raised_date", ""))
    } for q in queries]

    # Generate booklet to temp file
    tmp = tempfile.NamedTemporaryFile(
        suffix=".docx", delete=False,
        dir=os.path.join(os.path.dirname(__file__), "static"))
    tmp.close()
    fp = tmp.name

    generate_booklet(
        eng, tasks, cbt, rbt,
        queries_list,
        team, fp,
        firm_name=firm_name_str,
        firm_reg_no=firm_reg_no,
        audit_programs=audit_programs_list
    )

    log_action(g.firm_id, g.user["id"], "GENERATE_BOOKLET", "Engagement", eid,
               f"Generated booklet for engagement #{eid}")

    client_name = (eng.get("client_name") or "Engagement").replace(" ", "_")
    fy = eng.get("financial_year", "").replace("-", "_")
    download_name = f"Audit_Booklet_{client_name}_{fy}.docx"

    return send_file(
        fp,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )

# ═══════════════════════════════════════════════════════════════
#  AUDIT DAY LOGS
# ═══════════════════════════════════════════════════════════════

def _day_log_full(log):
    """Attach items to a day log dict."""
    log["items"] = qry(
        "SELECT * FROM audit_day_log_items WHERE log_id=%s ORDER BY item_type, sort_order",
        (log["id"],)
    )
    log["work"]        = [i for i in log["items"] if i["item_type"] == "work"]
    log["sampling"]    = [i for i in log["items"] if i["item_type"] == "sampling"]
    log["observations"]= [i for i in log["items"] if i["item_type"] == "observation"]
    log["pending"]     = [i for i in log["items"] if i["item_type"] == "pending"]
    return log


@app.route("/api/day-logs/")
@login_required
def list_day_logs():
    eid = request.args.get("engagement_id")
    if not eid:
        return jsonify({"detail": "engagement_id required"}), 400
    logs = qry(
        "SELECT * FROM audit_day_logs WHERE engagement_id=%s AND firm_id=%s ORDER BY log_date DESC",
        (eid, g.firm_id)
    )
    return jsonify(logs)


@app.route("/api/day-logs/<int:lid>")
@login_required
def get_day_log(lid):
    log = qry("SELECT * FROM audit_day_logs WHERE id=%s AND firm_id=%s",
              (lid, g.firm_id), one=True)
    if not log:
        return jsonify({"detail": "Not found"}), 404
    return jsonify(_day_log_full(log))


@app.route("/api/day-logs/", methods=["POST"])
@login_required
def save_day_log():
    d = request.get_json()
    eid       = d.get("engagement_id")
    log_date  = d.get("log_date")
    if not eid or not log_date:
        return jsonify({"detail": "engagement_id and log_date required"}), 400

    # Check engagement belongs to firm
    eng = qry("SELECT * FROM engagements WHERE id=%s AND firm_id=%s",
              (eid, g.firm_id), one=True)
    if not eng:
        return jsonify({"detail": "Engagement not found"}), 404

    # Check if log already exists for this date (upsert)
    existing = qry(
        "SELECT * FROM audit_day_logs WHERE engagement_id=%s AND firm_id=%s AND log_date=%s",
        (eid, g.firm_id, log_date), one=True
    )
    if existing and existing.get("is_locked"):
        return jsonify({"detail": "This log has been confirmed by client and is locked."}), 403

    if existing:
        lid = existing["id"]
        execute(
            "UPDATE audit_day_logs SET auditor_name=%s, next_day_plan=%s, updated_at=NOW() WHERE id=%s",
            (d.get("auditor_name", ""), d.get("next_day_plan", ""), lid)
        )
        execute("DELETE FROM audit_day_log_items WHERE log_id=%s", (lid,))
    else:
        lid = qry_id(
            "INSERT INTO audit_day_logs (firm_id, engagement_id, log_date, auditor_name, "
            "next_day_plan, created_by_id) VALUES (%s,%s,%s,%s,%s,%s)",
            (g.firm_id, eid, log_date, d.get("auditor_name", ""),
             d.get("next_day_plan", ""), g.user["id"])
        )

    # Insert items
    for idx, w in enumerate(d.get("work", [])):
        if w.get("area"):
            execute(
                "INSERT INTO audit_day_log_items (log_id,item_type,area,summary,staff,sort_order) "
                "VALUES (%s,'work',%s,%s,%s,%s)",
                (lid, w.get("area",""), w.get("summary",""),
                 ",".join(w.get("staff", [])), idx)
            )

    for idx, s in enumerate(d.get("sampling", [])):
        if s.get("area"):
            execute(
                "INSERT INTO audit_day_log_items "
                "(log_id,item_type,area,population,sample_size,materiality,filter_used,sort_order) "
                "VALUES (%s,'sampling',%s,%s,%s,%s,%s,%s)",
                (lid, s.get("area",""), s.get("population",""), s.get("sample_size",""),
                 s.get("materiality",""), s.get("filter_used",""), idx)
            )

    for idx, o in enumerate(d.get("observations", [])):
        if o.get("party") or o.get("query_text"):
            execute(
                "INSERT INTO audit_day_log_items "
                "(log_id,item_type,party,ref_no,amount,query_text,observation,mgmt_response,status,sort_order) "
                "VALUES (%s,'observation',%s,%s,%s,%s,%s,%s,%s,%s)",
                (lid, o.get("party",""), o.get("ref_no",""), o.get("amount",""),
                 o.get("query_text",""), o.get("observation",""),
                 o.get("mgmt_response",""), o.get("status","Open"), idx)
            )
            # Auto-create query in query sheet if status is Open
            if o.get("query_text") and o.get("status","Open") == "Open":
                max_sr = qry(
                    "SELECT COALESCE(MAX(sr_no),0) as m FROM queries WHERE engagement_id=%s",
                    (eid,), one=True
                )["m"]
                # Avoid duplicates — check if same query text already exists
                exists = qry(
                    "SELECT id FROM queries WHERE engagement_id=%s AND query_text=%s",
                    (eid, o.get("query_text","")), one=True
                )
                if not exists:
                    execute(
                        "INSERT INTO queries (firm_id,engagement_id,sr_no,query_text,"
                        "raised_by_id,task_reference,status) VALUES (%s,%s,%s,%s,%s,%s,'Open')",
                        (g.firm_id, eid, max_sr + 1, o.get("query_text",""),
                         g.user["id"], f"Day Log: {log_date}")
                    )

    for idx, p in enumerate(d.get("pending", [])):
        if p.get("doc_name"):
            execute(
                "INSERT INTO audit_day_log_items "
                "(log_id,item_type,doc_name,responsible,doc_status,sort_order) "
                "VALUES (%s,'pending',%s,%s,%s,%s)",
                (lid, p.get("doc_name",""), p.get("responsible",""),
                 p.get("doc_status","Pending"), idx)
            )
            # Auto-create task for pending document
            exists_task = qry(
                "SELECT id FROM tasks WHERE engagement_id=%s AND title=%s AND firm_id=%s",
                (eid, f"[Pending Doc] {p.get('doc_name','')}", g.firm_id), one=True
            )
            if not exists_task:
                qry_id(
                    "INSERT INTO tasks (firm_id,engagement_id,title,area,priority,status,created_by_id) "
                    "VALUES (%s,%s,%s,'Documentation','High','Pending',%s)",
                    (g.firm_id, eid,
                     f"[Pending Doc] {p.get('doc_name','')}",
                     g.user["id"])
                )

    log_action(g.firm_id, g.user["id"], "SAVE_DAY_LOG", "DayLog", lid,
               f"Day log saved for {log_date}", request.remote_addr)
    return jsonify({"id": lid, "message": "Day log saved"}), 201


@app.route("/api/day-logs/<int:lid>/confirm", methods=["POST"])
@login_required
def confirm_day_log(lid):
    """Client user confirms the day log — stamps and locks it."""
    log = qry("SELECT * FROM audit_day_logs WHERE id=%s AND firm_id=%s",
              (lid, g.firm_id), one=True)
    if not log:
        return jsonify({"detail": "Not found"}), 404
    if log.get("is_locked"):
        return jsonify({"detail": "Already confirmed and locked"}), 400
    if g.user["role"] != "Client":
        return jsonify({"detail": "Only a Client user can confirm a day log"}), 403

    execute(
        "UPDATE audit_day_logs SET is_confirmed=TRUE, confirmed_by=%s, "
        "confirmed_at=NOW(), confirmed_user_id=%s, is_locked=TRUE WHERE id=%s",
        (g.user["full_name"], g.user["id"], lid)
    )
    log_action(g.firm_id, g.user["id"], "CONFIRM_DAY_LOG", "DayLog", lid,
               f"Confirmed by {g.user['full_name']}", request.remote_addr)
    return jsonify({"message": "Day log confirmed and locked",
                    "confirmed_by": g.user["full_name"]})


@app.route("/api/day-logs/<int:lid>", methods=["DELETE"])
@require_role("Admin", "Team Leader")
def delete_day_log(lid):
    log = qry("SELECT * FROM audit_day_logs WHERE id=%s AND firm_id=%s",
              (lid, g.firm_id), one=True)
    if not log:
        return jsonify({"detail": "Not found"}), 404
    if log.get("is_locked"):
        return jsonify({"detail": "Cannot delete a confirmed log"}), 403
    execute("DELETE FROM audit_day_log_items WHERE log_id=%s", (lid,))
    execute("DELETE FROM audit_day_logs WHERE id=%s", (lid,))
    return jsonify({"message": "Deleted"})

# ═══════════════════════════════════════════════════════════════
#  HEALTH CHECK
# ═══════════════════════════════════════════════════════════════
@app.route("/api/health")
def health():
    return jsonify({"status":"ok","version":"saas-v1"})

# ═══════════════════════════════════════════════════════════════
#  TOOL: TAX AUDIT (SEC 44AB) ANALYZER
# ═══════════════════════════════════════════════════════════════
@app.route("/api/tools/tax-audit/analyze", methods=["POST"])
@login_required
def tax_audit_analyze():
    try:
        import pandas as pd
    except ImportError:
        return jsonify({"detail": "pandas not installed. Contact support."}), 500
    try:
        import openpyxl
    except ImportError:
        return jsonify({"detail": "openpyxl not installed. Contact support."}), 500

    software    = request.form.get("software", "Tally")
    nature      = request.form.get("nature", "Business")
    presumptive = request.form.get("presumptive", "No")
    try:
        turnover = float(request.form.get("turnover", "0").replace(",", ""))
    except Exception:
        return jsonify({"detail": "Invalid turnover value"}), 400

    cash_files = request.files.getlist("cash_files")
    bank_files = request.files.getlist("bank_files")
    od_files   = request.files.getlist("od_files")

    if not cash_files and not bank_files and not od_files:
        return jsonify({"detail": "Please upload at least one ledger file."}), 400

    def clean_numeric(val):
        if val is None: return 0.0
        try:
            if pd.isna(val): return 0.0
        except Exception: pass
        if isinstance(val, (int, float)): return float(val)
        val = str(val).replace(",","").replace("₹","").replace("Rs.","").replace(" ","").strip()
        try: return float(val)
        except: return 0.0

    def check_ignore(row_vals):
        keywords = ["opening balance","closing balance","brought forward","carried forward",
                    "balance b/d","balance c/d","balance b/f","balance c/f","total","grand total"]
        for v in row_vals:
            vs = str(v).lower().strip()
            if any(k in vs for k in keywords): return True
        return False

    def read_file(f, source_type):
        import io
        fname = f.filename
        try:
            raw = f.read()
            buf = io.BytesIO(raw)
            if fname.lower().endswith(".csv"):
                df_raw = pd.read_csv(buf, header=None)
            else:
                df_raw = pd.read_excel(buf, header=None)
        except Exception as e:
            return None, str(e)

        try:
            header_idx = -1
            for i, row in df_raw.head(200).iterrows():
                row_str = " ".join(str(x).lower() for x in row.values)
                if any(k in row_str for k in ["date","dt","particulars","narration"]):
                    header_idx = i; break

            if header_idx != -1:
                df = df_raw.iloc[header_idx:].reset_index(drop=True)
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)
                df.columns = [str(c).strip().lower() for c in df.columns]
            else:
                df = df_raw.copy()
                df.columns = [f"col_{i}" for i in range(len(df.columns))]

            col_map = {}
            for c in df.columns:
                if not col_map.get("Date") and any(x in c for x in ["date","dt","col_0"]):
                    col_map["Date"] = c
                elif not col_map.get("Particulars") and any(x in c for x in ["particulars","narration","desc","detail","col_2"]):
                    col_map["Particulars"] = c
                elif not col_map.get("VchType") and any(x in c for x in ["vch","type","voucher","col_3"]):
                    col_map["VchType"] = c
                elif not col_map.get("Debit") and any(x in c for x in ["debit","dr","receipt","deposit","col_5"]):
                    col_map["Debit"] = c
                elif not col_map.get("Credit") and any(x in c for x in ["credit","cr","payment","withdrawal","col_6"]):
                    col_map["Credit"] = c

            if "Date" not in col_map or "Particulars" not in col_map:
                return None, f"Could not identify columns in {fname}"

            df["_date"] = pd.to_datetime(df[col_map["Date"]], dayfirst=True, errors="coerce")
            df["_ignore"] = df.apply(lambda r: check_ignore(r.values), axis=1)
            merged = []
            for i, row in df.iterrows():
                if pd.isna(row["_date"]) and not row["_ignore"]: continue
                narr = str(row.get(col_map["Particulars"], "")).strip()
                vch  = str(row.get(col_map.get("VchType",""), "")).strip()
                dr   = clean_numeric(row.get(col_map.get("Debit",""), 0))
                cr   = clean_numeric(row.get(col_map.get("Credit",""), 0))
                ign  = row["_ignore"]
                if dr > 0 or cr > 0:
                    merged.append({
                        "Date": row["_date"], "Narration": narr,
                        "Debit": dr, "Credit": cr,
                        "Source": source_type, "File": fname,
                        "VchType": vch, "Tag": "IGNORE_ROW" if ign else "NORMAL"
                    })
            return pd.DataFrame(merged), None
        except Exception as e:
            return None, str(e)

    all_dfs = []; failed = []
    for flist, stype in [(cash_files,"CASH"),(bank_files,"BANK"),(od_files,"BANK_OD")]:
        for f in flist:
            if not f or not f.filename: continue
            df, err = read_file(f, stype)
            if df is not None and not df.empty:
                all_dfs.append(df)
            else:
                failed.append(f.filename + (f" ({err})" if err else ""))

    if not all_dfs:
        return jsonify({"detail": f"No valid data found. Failed: {', '.join(failed)}"}), 400

    df = pd.concat(all_dfs, ignore_index=True).sort_values("Date").reset_index(drop=True)
    df = df[df["Tag"] != "IGNORE_ROW"].copy().reset_index(drop=True)

    # Tag contra / cheque return entries
    return_kw = ["return","bounce","dishonour","reject","reversal","unpaid"]
    for i, row in df.iterrows():
        if df.at[i,"Tag"] != "NORMAL": continue
        vl = str(row.get("VchType","")).lower()
        nl = str(row["Narration"]).lower()
        if "contra" in vl or "ctra" in vl:
            df.at[i,"Tag"] = "CONTRA"
        elif any(k in nl for k in return_kw) and row["Source"] in ["BANK","BANK_OD"]:
            df.at[i,"Tag"] = "CHEQUE_RETURN"

    bank = df[df["Source"].isin(["BANK","BANK_OD"])]
    cash = df[df["Source"]=="CASH"]

    gross_bank_rx   = float(bank[bank["Debit"]>0]["Debit"].sum())
    bank_contra_rx  = float(bank[(bank["Tag"]=="CONTRA")&(bank["Debit"]>0)]["Debit"].sum())
    bank_inter_rx   = 0.0
    bank_return_rx  = float(bank[(bank["Tag"]=="CHEQUE_RETURN")&(bank["Debit"]>0)]["Debit"].sum())
    net_bank_rx     = gross_bank_rx - bank_contra_rx - bank_inter_rx - bank_return_rx

    gross_bank_pmt  = float(bank[bank["Credit"]>0]["Credit"].sum())
    bank_contra_pmt = float(bank[(bank["Tag"]=="CONTRA")&(bank["Credit"]>0)]["Credit"].sum())
    bank_inter_pmt  = 0.0
    bank_return_pmt = float(bank[(bank["Tag"]=="CHEQUE_RETURN")&(bank["Credit"]>0)]["Credit"].sum())
    net_bank_pmt    = gross_bank_pmt - bank_contra_pmt - bank_inter_pmt - bank_return_pmt

    gross_cash_rx   = float(cash[cash["Debit"]>0]["Debit"].sum())
    cash_contra_rx  = float(cash[(cash["Tag"]=="CONTRA")&(cash["Debit"]>0)]["Debit"].sum())
    net_cash_rx     = gross_cash_rx - cash_contra_rx

    gross_cash_pmt  = float(cash[cash["Credit"]>0]["Credit"].sum())
    cash_contra_pmt = float(cash[(cash["Tag"]=="CONTRA")&(cash["Credit"]>0)]["Credit"].sum())
    net_cash_pmt    = gross_cash_pmt - cash_contra_pmt

    total_rx  = net_bank_rx + net_cash_rx
    total_pmt = net_bank_pmt + net_cash_pmt
    cash_rx_pct  = round((net_cash_rx / total_rx * 100)  if total_rx  > 0 else 0.0, 2)
    cash_pmt_pct = round((net_cash_pmt / total_pmt * 100) if total_pmt > 0 else 0.0, 2)

    # Audit decision
    audit_status = "Not Applicable"; section = "N/A"; reason = "Below threshold limits"
    if nature == "Profession":
        if turnover > 5000000:
            audit_status = "Applicable"; section = "Sec 44AB(b)"
            reason = "Gross receipts from profession > ₹50 Lakhs"
    else:
        if presumptive == "Yes (Opted Out)":
            audit_status = "Applicable"; section = "Sec 44AB(e)"
            reason = "Opted out of presumptive taxation (Sec 44AD(4))"
        elif turnover > 100000000:
            audit_status = "Applicable"; section = "Sec 44AB(a)"
            reason = "Total Turnover > ₹10 Crores"
        elif turnover > 10000000:
            if cash_rx_pct <= 5.0 and cash_pmt_pct <= 5.0:
                audit_status = "Not Applicable"; section = "Proviso to Sec 44AB(a)"
                reason = "Turnover > ₹1 Cr but Cash transactions ≤ 5%"
            else:
                audit_status = "Applicable"; section = "Sec 44AB(a)"
                reason = f"Turnover > ₹1 Cr AND Cash Rx ({cash_rx_pct}%) or Pmt ({cash_pmt_pct}%) > 5%"

    # Ledger-wise working
    ledger_working = []
    for (src, fn), fdf in df.groupby(["Source","File"]):
        ledger_working.append({
            "book_type": src, "file": fn,
            "gross_rx":       float(fdf[fdf["Debit"]>0]["Debit"].sum()),
            "less_contra_rx": float(fdf[(fdf["Tag"]=="CONTRA")&(fdf["Debit"]>0)]["Debit"].sum()),
            "less_inter_rx":  0.0,
            "less_return_rx": float(fdf[(fdf["Tag"]=="CHEQUE_RETURN")&(fdf["Debit"]>0)]["Debit"].sum()),
            "net_rx":         float(fdf[(fdf["Tag"]=="NORMAL")&(fdf["Debit"]>0)]["Debit"].sum()),
            "gross_pmt":      float(fdf[fdf["Credit"]>0]["Credit"].sum()),
            "net_pmt":        float(fdf[(fdf["Tag"]=="NORMAL")&(fdf["Credit"]>0)]["Credit"].sum()),
        })

    # Flags: 40A(3) and 269ST
    flags = []
    normal_cash = df[(df["Tag"]=="NORMAL")&(df["Source"]=="CASH")]
    for _, r in normal_cash.iterrows():
        if r["Credit"] > 10000:
            flags.append({"date": str(r["Date"])[:10], "narration": r["Narration"],
                          "amount": float(r["Credit"]), "warning": "Sec 40A(3) — Cash Payment > ₹10,000"})
        if r["Debit"] >= 200000:
            flags.append({"date": str(r["Date"])[:10], "narration": r["Narration"],
                          "amount": float(r["Debit"]), "warning": "Sec 269ST — Cash Receipt ≥ ₹2,00,000"})

    return jsonify({
        "summary": {
            "gross_bank_rx": gross_bank_rx, "bank_contra_rx": bank_contra_rx,
            "bank_inter_rx": bank_inter_rx, "bank_return_rx": bank_return_rx,
            "net_bank_rx": net_bank_rx,
            "gross_bank_pmt": gross_bank_pmt, "bank_contra_pmt": bank_contra_pmt,
            "bank_inter_pmt": bank_inter_pmt, "bank_return_pmt": bank_return_pmt,
            "net_bank_pmt": net_bank_pmt,
            "gross_cash_rx": gross_cash_rx, "cash_contra_rx": cash_contra_rx,
            "net_cash_rx": net_cash_rx,
            "gross_cash_pmt": gross_cash_pmt, "cash_contra_pmt": cash_contra_pmt,
            "net_cash_pmt": net_cash_pmt,
            "total_net_rx": total_rx, "total_net_pmt": total_pmt,
            "cash_rx_pct": cash_rx_pct, "cash_pmt_pct": cash_pmt_pct,
            "audit_status": audit_status, "section": section, "reason": reason,
            "failed_files": failed
        },
        "ledger_working": ledger_working,
        "flags": flags,
        "matched_pairs": []
    })


@app.route("/api/tools/tax-audit/export-excel", methods=["POST"])
@login_required
def tax_audit_export_excel():
    import io, os, tempfile
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"detail": "openpyxl not installed"}), 500

    data   = request.get_json(silent=True) or {}
    s      = data.get("summary", {})
    lw     = data.get("ledger_working", [])
    client = data.get("client", {})

    def inr(v):
        try: return f"₹{float(v):,.2f}"
        except: return "₹0.00"

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "44AB Working"
    navy_fill = PatternFill("solid", fgColor="003366")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    reg_font  = Font(name="Arial", size=10)
    thin      = Side(style="thin", color="D1DAEA")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, val, bg="FFFFFF", bold=False, align="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = Font(name="Arial", bold=bold, size=10)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = bdr
        return c

    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="Tax Audit (Sec 44AB) Working Report")
    t.font = Font(name="Arial", bold=True, size=14, color="003366")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    info = [("Assessee", client.get("name","")), ("PAN", client.get("pan","")),
            ("A.Y.", client.get("ay","")), ("Nature", client.get("nature","")),
            ("Turnover", inr(client.get("turnover",0)))]
    for i, (k, v) in enumerate(info, 2):
        cell(i, 1, k, bg="F0F4F8", bold=True)
        cell(i, 2, v)

    r = 8
    ws.merge_cells(f"A{r}:B{r}")
    h = ws.cell(row=r, column=1, value="44AB Working Analysis")
    h.font = hdr_font; h.fill = navy_fill
    h.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    rows_data = [
        ("Bank Gross Receipts", s.get("gross_bank_rx",0)),
        ("  Less: Contra (Receipts)", s.get("bank_contra_rx",0)),
        ("  Less: Cheque Returns", s.get("bank_return_rx",0)),
        ("Net Bank Receipts (A)", s.get("net_bank_rx",0), True),
        ("Bank Gross Payments", s.get("gross_bank_pmt",0)),
        ("  Less: Contra (Payments)", s.get("bank_contra_pmt",0)),
        ("Net Bank Payments (B)", s.get("net_bank_pmt",0), True),
        ("Gross Cash Receipts", s.get("gross_cash_rx",0)),
        ("  Less: Contra", s.get("cash_contra_rx",0)),
        ("Net Cash Receipts (C)", s.get("net_cash_rx",0), True),
        ("Gross Cash Payments", s.get("gross_cash_pmt",0)),
        ("Net Cash Payments (D)", s.get("net_cash_pmt",0), True),
        ("Total Net Receipts (A+C)", s.get("total_net_rx",0), True),
        ("Total Net Payments (B+D)", s.get("total_net_pmt",0), True),
        (f"Cash Receipt %", f"{s.get('cash_rx_pct',0):.2f}%", True),
        (f"Cash Payment %", f"{s.get('cash_pmt_pct',0):.2f}%", True),
    ]
    for rd in rows_data:
        bold = len(rd) > 2 and rd[2]
        bg = "EBF4FF" if bold else "FFFFFF"
        cell(r, 1, rd[0], bg=bg, bold=bold)
        cell(r, 2, inr(rd[1]) if isinstance(rd[1], (int, float)) else rd[1],
             bg=bg, bold=bold, align="right")
        r += 1

    r += 1
    status = s.get("audit_status","")
    status_bg = "FFF0F0" if status == "Applicable" else "E8F5EE"
    ws.merge_cells(f"A{r}:B{r}")
    sc = ws.cell(row=r, column=1,
                 value=f"AUDIT STATUS: {status} — {s.get('section','')} — {s.get('reason','')}")
    sc.font = Font(name="Arial", bold=True, size=11,
                   color="C53030" if status == "Applicable" else "1E7E44")
    sc.fill = PatternFill("solid", fgColor=status_bg.replace("#",""))
    sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 30

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False,
          dir=os.path.join(os.path.dirname(__file__), "static"))
    tmp.close()
    wb.save(tmp.name)
    return send_file(tmp.name, as_attachment=True,
                     download_name="Tax_Audit_44AB_Report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════
#  FILE UPLOADS (Cloudflare R2 or local fallback)
# ═══════════════════════════════════════════════════════════════
def _upload_to_r2(file_bytes, key, content_type):
    """Upload bytes to Cloudflare R2. Returns public URL."""
    if not R2_ACCESS_KEY:
        return None   # Fall back to local storage
    import boto3
    endpoint = f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com"
    s3 = boto3.client("s3",
        endpoint_url=endpoint,
        aws_access_key_id=R2_ACCESS_KEY,
        aws_secret_access_key=R2_SECRET_KEY,
        region_name="auto"
    )
    s3.put_object(Bucket=R2_BUCKET, Key=key, Body=file_bytes, ContentType=content_type)
    return key

@app.route("/api/files/")
@login_required
def list_files():
    tid = request.args.get("task_id")
    rows = qry(
        "SELECT fu.*, u.full_name as uploaded_by_name FROM file_uploads fu "
        "JOIN users u ON fu.uploaded_by_id=u.id "
        "WHERE fu.task_id=%s AND fu.firm_id=%s ORDER BY fu.uploaded_at DESC",
        (tid, g.firm_id))
    return jsonify(rows)

@app.route("/api/files/upload", methods=["POST"])
@login_required
def upload_file():
    tid  = request.form.get("task_id")
    f    = request.files.get("file")
    if not f: return jsonify({"detail":"No file"}), 400
    task = qry("SELECT * FROM tasks WHERE id=%s AND firm_id=%s", (tid, g.firm_id), one=True)
    if not task: return jsonify({"detail":"Task not found"}), 404
    raw = f.read()
    key = f"firm_{g.firm_id}/eng_{task['engagement_id']}/task_{tid}/{uuid.uuid4().hex[:8]}_{f.filename}"
    _upload_to_r2(raw, key, f.content_type or "application/octet-stream")
    fid = qry_id(
        "INSERT INTO file_uploads (firm_id,task_id,filename,original_filename,file_key,"
        "file_size,mime_type,uploaded_by_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
        (g.firm_id, tid, f.filename, f.filename, key, len(raw), f.content_type, g.user["id"])
    )
    return jsonify({"id": fid, "filename": f.filename}), 201

@app.route("/api/files/<int:fid>", methods=["DELETE"])
@login_required
def delete_file(fid):
    execute("DELETE FROM file_uploads WHERE id=%s AND firm_id=%s", (fid, g.firm_id))
    return jsonify({"message":"Deleted"})

# ═══════════════════════════════════════════════════════════════
#  STARTUP
# ═══════════════════════════════════════════════════════════════
with app.app_context():
    if DATABASE_URL:
        init_db()
    else:
        print("[WARN] DATABASE_URL not set — skipping DB init")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT, debug=False)
